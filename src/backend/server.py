# -*- coding: utf-8 -*-
"""
Dit script fungeert als de backend-server voor de Excel-PDF-automaton.
Het is aangepast voor productiehosting met Docker en Gunicorn.
"""

# --- Imports ---
import os
import re
import json
import time
import uuid
import traceback
import io  # <-- Nieuwe import voor in-memory bestandsverwerking
from datetime import datetime

import pandas as pd
import pdfplumber
import pytesseract
import requests
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pdf2image import convert_from_bytes # <-- Aangepast van convert_from_path
from openpyxl.styles import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
try:
    from PIL import Image
except ImportError:
    import Image

# --- Configuratie ---
# Voor productie is het beter om dit via een omgevingsvariabele te doen.
# Render.com laat ons dit makkelijk instellen.
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "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")
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL_NAME = "deepseek/deepseek-r1:free"
print("Configuratie geladen.")


# --- Functiedefinities (Aangepast voor In-Memory Verwerking) ---

def extract_text_from_pdf_with_ocr(pdf_filename, pdf_bytes, min_text_length=100):
    """
    Extraheert tekst uit PDF-bytes. Schakelt over naar OCR indien nodig.
    """
    text = ""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf: # <-- Leest uit bytes
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text: text += page_text + "\n"
        print(f"Methode: Digitale extractie geslaagd voor {pdf_filename}.")
    except Exception as e:
        print(f"Fout tijdens digitale extractie voor {pdf_filename}: {e}. Probeert nu OCR.")
        text = ""

    if len(text.strip()) >= min_text_length:
        return text

    print(f"Digitale extractie voor {pdf_filename} leverde te weinig tekst op. Overschakelen naar OCR...")
    try:
        images = convert_from_bytes(pdf_bytes) # <-- Leest uit bytes
        ocr_text = ""
        for i, image in enumerate(images):
            print(f"  Verwerken van pagina {i+1} met OCR...")
            ocr_text += pytesseract.image_to_string(image, lang='nld') + "\n"
        print(f"Methode: OCR-extractie (Tesseract) geslaagd voor {pdf_filename}.")
        return ocr_text
    except Exception as e:
        print(f"Fout tijdens OCR-extractie voor {pdf_filename}: {e}")
        return None

def extract_percentage_sentences(text):
    if not text:
        return []
    text = text.replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text)
    sentences = re.split(r'(?<=[.?!])\s+', text)
    keywords = ['loon', 'salaris', 'cao', 'verhoging', 'stijging', 'toeslag', 'schaalsalarissen']
    percentage_pattern = re.compile(r'\d[\d.,]*\s?%')
    relevant_sentences = []
    for sentence in sentences:
        if not sentence: continue
        sentence_lower = sentence.lower()
        has_keyword = any(keyword in sentence_lower for keyword in keywords)
        has_percentage = percentage_pattern.search(sentence_lower)
        if has_keyword and has_percentage:
            relevant_sentences.append(sentence.strip())
    return relevant_sentences

def classify_with_deepseek(sentence, api_key, api_url, model_name, max_retries=3, delay=2):
    headers = {
        "Authorization": f"Bearer {api_key}", "Content-Type": "application/json",
        "HTTP-Referer": "https://localhost/sakkal-cao-analyzer", "X-Title": "Team Sakkal CAO Analyzer"
    }
    system_prompt = """Je bent een expert in het analyseren van Nederlandse CAO-teksten. Je taak is om **alleen** concrete, definitieve loonstijgingen te vinden, te extraheren en te categoriseren.

**ZEER BELANGRIJKE REGELS OM TE VOLGEN:**
- **NEGEER VOLLEDIG** alle zinnen die onderdeel zijn van een voorbeeld, berekening of hypothese.
- Zoek naar sleutelwoorden zoals: **"Voorbeeld:", "Rekenvoorbeeld", "Stel dat", "Als ... dan", "Berekening"**. Als je zo'n sleutelwoord ziet, is de zin bijna altijd een voorbeeld en moet je een lege `verhogingen` lijst teruggeven.
- Negeer ook verhogingen die voorwaardelijk zijn ('tenzij', 'indien') of die alleen een vergelijking maken met het verleden.
- Focus **alleen** op definitieve, vastgelegde collectieve loonsverhogingen voor de toekomst.

**Jouw taak:**
Analyseer de zin en geef een JSON-object terug. Het object moet een lijst `verhogingen` bevatten. Voor **elke** gevonden loonstijging, maak een object met VIER sleutels:
1. `datum`: De ingangsdatum (formaat: "DD/MM/YYYY").
2. `percentage`: Het percentage (als een getal, bv. 3.5).
3. `categorie`: Classificeer het type verhoging. Kies uit: "standaard", "verlofdag_omzetting", "dienstjaren_toeslag", "WML_koppeling", "anders".
4. `uitleg`: Een korte toelichting op je keuze.

**BELANGRIJK:** Je antwoord MOET **uitsluitend** een enkel, geldig JSON-object zijn met de sleutel 'verhogingen'. Geen extra tekst."""
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": "De salarissen stijgen met 2% op 01-01-2025 en met nog eens 3% op 01-07-2025."},
        {"role": "assistant", "content": '''{"verhogingen": [{"datum": "01/01/2025", "percentage": 2.0, "categorie": "standaard", "uitleg": "Een standaard collectieve verhoging."}, {"datum": "01/07/2025", "percentage": 3.0, "categorie": "standaard", "uitleg": "Een tweede standaard collectieve verhoging."}]}'''},
        {"role": "user", "content": "Het minimumloon wordt per 1/1/26 berekend door 3,85% bovenop het WML te tellen."},
        {"role": "assistant", "content": '''{"verhogingen": [{"datum": "01/01/2026", "percentage": 3.85, "categorie": "WML_koppeling", "uitleg": "De verhoging is direct gekoppeld aan het WML."}]}'''},
        {"role": "user", "content": "Voorbeeld: Als het bruto uursalaris met 3,88% stijgt, dan..."},
        {"role": "assistant", "content": '''{"verhogingen": [], "uitleg": "De zin begint met 'Voorbeeld:'."}'''},
        {"role": "user", "content": sentence}
    ]
    data = { "model": MODEL_NAME, "messages": messages, "response_format": { "type": "json_object" }, "max_tokens": 500, "temperature": 0.0 }
    for attempt in range(max_retries):
        try:
            response = requests.post(api_url, json=data, headers=headers, timeout=45)
            response.raise_for_status()
            full_response_json = response.json()
            message_content_str = full_response_json['choices'][0]['message']['content']
            json_start = message_content_str.find('{')
            json_end = message_content_str.rfind('}')
            if json_start != -1 and json_end != -1:
                return json.loads(message_content_str[json_start:json_end+1])
            return None
        except Exception as e:
            print(f"Fout in classify_with_deepseek: {e}")
            if attempt < max_retries - 1: time.sleep(delay)
    return None

def analyze_pdfs(uploaded_files):
    """
    Orchestreert het volledige analyseproces voor een lijst van geüploade bestanden in-memory.
    """
    final_results = {}
    for file_data in uploaded_files:
        filename = file_data['name']
        content = file_data['content']
        print(f'--- Start analyse voor: {filename} ---')
        
        extracted_text = extract_text_from_pdf_with_ocr(filename, content)
        if not extracted_text:
            final_results[filename] = {"error": "Tekstextractie mislukt", "verhogingen": []}
            continue
        
        sentences = extract_percentage_sentences(extracted_text)
        if not sentences:
            final_results[filename] = {"error": "Geen relevante zinnen gevonden", "verhogingen": []}
            print(f"  -> ❌ Geen relevante zinnen gevonden in {filename}.")
            continue

        print(f"  -> ✅ {len(sentences)} potentieel relevante zin(nen) gevonden.")
        all_found_increases = []
        for i, sentence in enumerate(sentences):
            print(f"  Verwerken zin {i+1}/{len(sentences)}...")
            result_json = classify_with_deepseek(sentence, OPENROUTER_API_KEY, OPENROUTER_API_URL, MODEL_NAME)
            if result_json and 'verhogingen' in result_json and isinstance(result_json['verhogingen'], list):
                if len(result_json['verhogingen']) > 0:
                    print(f"    -> ✅ {len(result_json['verhogingen'])} verhoging(en) gevonden in zin.")
                    all_found_increases.extend(result_json['verhogingen'])
                else:
                    print("    -> ❌ Geen concrete verhoging geclassificeerd in zin.")
            else:
                 print("    -> ❓ Fout bij classificeren of geen verhogingen gevonden.")
            time.sleep(1.5)
        
        final_results[filename] = {"verhogingen": all_found_increases}
        print(f'--- Analyse voor {filename} voltooid ---')
    return final_results

def create_excel_summary(analysis_results):
    """
    Maakt een Excel-bestand in-memory en retourneert het als een BytesIO-buffer.
    """
    output_buffer = io.BytesIO()
    COLOR_MAP = {
        "verlofdag_omzetting": "FFC000", "dienstjaren_toeslag": "0070C0",
        "WML_koppeling": "00B050", "anders": "7030A0", "standaard": "000000"
    }
    RICH_TEXT_FONT_MAP = {cat: InlineFont(color=color) for cat, color in COLOR_MAP.items()}
    CELL_STYLE_FONT_MAP = {cat: Font(color=color) for cat, color in COLOR_MAP.items()}
    DEFAULT_RICH_TEXT_FONT = RICH_TEXT_FONT_MAP["standaard"]
    LEGEND = {
        "standaard": "Standaard loonsverhoging.", "WML_koppeling": "Gekoppeld aan WML.",
        "verlofdag_omzetting": "Omzetting van verlofdagen.", "dienstjaren_toeslag": "O.b.v. dienstjaren.",
        "anders": "Andere specifieke verhoging."
    }

    grouped_results, max_dates = {}, 0
    for filename, data in analysis_results.items():
        by_date = {}
        for inc in data.get('verhogingen', []):
            date_str = inc.get('datum')
            if not date_str or not isinstance(inc.get('percentage'), (int, float)): continue
            try:
                date_key = datetime.strptime(date_str, '%d/%m/%Y')
                if date_key not in by_date: by_date[date_key] = []
                by_date[date_key].append(inc)
            except (ValueError, TypeError): continue
        sorted_dates = sorted(by_date.keys())
        grouped_results[filename] = [(d.strftime('%d/%m/%Y'), by_date[d]) for d in sorted_dates]
        max_dates = max(max_dates, len(sorted_dates))

    processed_data = []
    for filename, date_groups in grouped_results.items():
        row = {'Bestandsnaam': filename}
        for i, (date, increases) in enumerate(date_groups):
            row[f'{i+1}e datum'] = date
            row[f'{i+1}e percentages'] = " / ".join([f"{inc.get('percentage', 0):.2f}%".replace('.', ',') for inc in increases])
        processed_data.append(row)

    df = pd.DataFrame()
    if processed_data:
        headers = ['Bestandsnaam'] + [item for i in range(1, max_dates + 1) for item in (f'{i}e datum', f'{i}e percentages')]
        df = pd.DataFrame(processed_data, columns=headers).fillna('')

    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Samenvatting')
        ws = writer.sheets['Samenvatting']
        header_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
        for row_idx, filename in enumerate(df['Bestandsnaam'], start=2):
            for i, (date, increases) in enumerate(grouped_results.get(filename, [])):
                col_idx = header_map.get(f'{i+1}e percentages')
                if not col_idx: continue
                cell = ws.cell(row=row_idx, column=col_idx)
                payload = []
                for j, inc in enumerate(increases):
                    if j > 0 and payload: payload.append(TextBlock(DEFAULT_RICH_TEXT_FONT, " / "))
                    p_str = f"{inc.get('percentage', 0):.2f}%".replace('.', ',')
                    font = RICH_TEXT_FONT_MAP.get(inc.get('categorie', 'standaard'), DEFAULT_RICH_TEXT_FONT)
                    payload.append(TextBlock(font, p_str))
                if payload: cell.value = CellRichText(payload)
        for i, col_cells in enumerate(ws.columns):
            length = max(len(str(cell.value)) if not isinstance(cell.value, CellRichText) else len("".join(b.text for b in cell.value)) for cell in col_cells)
            ws.column_dimensions[get_column_letter(i + 1)].width = length + 4
        if not df.empty:
            legend_col = len(df.columns) + 2
            ws.cell(row=1, column=legend_col, value="LEGENDA").font = Font(bold=True)
            for r, (cat, desc) in enumerate(LEGEND.items(), start=2):
                font_style = CELL_STYLE_FONT_MAP.get(cat, CELL_STYLE_FONT_MAP['standaard'])
                ws.cell(row=r, column=legend_col, value=desc).font = font_style
            ws.column_dimensions[get_column_letter(legend_col)].width = max(len(d) for d in LEGEND.values()) + 4

    print(f'Excel-bestand succesvol aangemaakt in-memory.')
    output_buffer.seek(0)
    return output_buffer


# --- Flask Webserver ---
app = Flask(__name__)
CORS(app)

@app.route('/api/process', methods=['POST'])
def process_uploaded_pdfs():
    print("\n[SERVER LOG] Received new request for /api/process")
    if 'files' not in request.files:
        return jsonify({"error": "Geen bestanden meegegeven"}), 400
    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        return jsonify({"error": "Geen bestanden geselecteerd"}), 400

    uploaded_files_data = []
    for f in files:
        filename = secure_filename(f.filename)
        uploaded_files_data.append({
            "name": filename,
            "content": f.read() # Lees de inhoud van het bestand in bytes
        })

    try:
        analysis_results = analyze_pdfs(uploaded_files_data)
        if not analysis_results or all(not v.get('verhogingen') for v in analysis_results.values()):
            return jsonify({"error": "Analyse heeft geen bruikbare loonsverhogingen opgeleverd."}), 500

        excel_buffer = create_excel_summary(analysis_results)
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='cao_samenvatting.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Onverwachte serverfout: {str(e)}"}), 500

# De 'if __name__ == '__main__':' blok is niet nodig wanneer Gunicorn wordt gebruikt,
# maar kan hier blijven staan voor eventuele lokale tests. Gunicorn start de 'app'
# variabele direct.
# if __name__ == '__main__':
#    app.run(host='127.0.0.1', port=5001, debug=True) 